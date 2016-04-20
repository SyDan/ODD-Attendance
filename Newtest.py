from tkinter import *
from tkinter import ttk
from openpyxl import *
from tkinter import messagebox
from tkinter import filedialog
from PIL import Image, ImageTk

#import time
import os
#from openpyxl.compat import range
import datetime
#from openpyxl.cell import get_column_letter
from openpyxl.styles import PatternFill#, Color, Font, Border
from twilio.rest import TwilioRestClient

from googleapiclient.http import MediaIoBaseDownload   #imports of Google Drive Upload download scripts
from googleapiclient.discovery import build
from httplib2 import Http
from oauth2client import file, client, tools


redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
yellowFill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
whiteFill = PatternFill()
Roll = []

def rowcol():
    global trow, tcol, tot_stud, tot_class
    for i in range(6,100):
        if ws.cell(row=i , column=2).value==None:
            trow=(i-1)
            break
    for i in range(8,60):
        if ws.cell(row=7, column=i).value==None:
            tcol=(i-1)
            break
    tot_stud = trow-5
    tot_class = tcol-7

def rawt():                                         #Inserts all roll no. in Roll[]
    for i in range(6, trow+1):
        Roll.append(ws.cell(row=i, column=2).value)

def parti_Atten(sl):                                #INserts all attendace of a particular student in list
    real_sl=sl+5
    xroll=[]
    for i in range(8, tcol+1):
        xroll.append(ws.cell(row=real_sl, column=i).value)
    return xroll

def sum_par():                                      #INserts total attendance and % attendance in excel sheets
    for i in range(1, trow-4):
        Aroll = parti_Atten(i)
        ws.cell(row=i+5, column=6).value=sum(Aroll)
        ws.cell(row=i+5, column=7).value=round((sum(Aroll)/len(Aroll)*100),2)
def sum_par1(*args):
    sum_par()
    wb.save('{}_ODD.xlsx'.format(filen))
    messagebox.showinfo(message='Total Attendance and Attendance % has been UPDATED as a new file named\n\n{}_ODD.xlsx'.format(filen), icon='info', title='Sheet Updated')
def cur_parti_Atte(sl):                             #TAkes last 5 attendace of a particular student
    xroll = []
    for i in range(tcol-4, tcol+1):
        xroll.append(ws.cell(row=sl+5, column=i).value)
    return xroll

def tim():                                         #Inserts time for new attendance
        x = str(datetime.datetime.now())
        ws.cell(row=5, column=tcol+1).value='{}/{}'.format(x[8:10], x[5:7])


def debar():                                       #Inserts debarred information in a excel sheet
    p=messagebox.askyesno(message='Debar List will be updated strictly according to the rules:\n1. If Attendance < 75%\n2. If 5 consecutive Absence\nWould you like to UPDATE in a new file named\n\n{}_ODD.xlsx?'.format(filen), icon='question', title='Debar List Updation')
    if int(p)==1:
        for i in range(1, trow-4):
            ws.cell(row=i+5, column=5).fill=whiteFill
            ws.cell(row=i+5, column=4).fill=whiteFill
            Aroll=parti_Atten(i)
            sAroll=str(Aroll)
            if round((sum(Aroll)/len(Aroll)*100),2)<75:
                ws.cell(row=i+5, column=4).value=1
                ws.cell(row=i+5, column=4).fill=redFill
            if '0, 0, 0, 0, 0' in sAroll:
                ws.cell(row=i+5, column=5).value=1
                ws.cell(row=i+5, column=5).fill=yellowFill
        wb.save('{}_ODD.xlsx'.format(filen))






#rowcol()
#rawt()
#sum_per()
#time_now=tim()
#debar()
#wb.save(filename.get())







root=Tk()
root.title('ODD Attendance')
root.grid_columnconfigure(0, weight=1)
root.grid_rowconfigure(0, weight=1)
nb = ttk.Notebook(root)#, padding='10 10'
nb.grid(row=0, column=0, sticky='nsew')
nb.columnconfigure(0, weight=1)
nb.rowconfigure(0, weight=1)

roll_l1 = StringVar()
roll_l2 = StringVar()
y_warn_text = StringVar()
warn3=StringVar()
warn75=StringVar()
per = StringVar()
counter=0
roll_var = StringVar()
done_debar_low = StringVar()
done_debar_con = StringVar()
done_abs = StringVar()
done_warn = StringVar()
done_absdebar = StringVar()
att = StringVar()

open_f = ttk.Frame(nb, borderwidth=10, relief='solid')#padding='5 5 5 5
open_f.grid(row=0, column=0, sticky='nsew')
open_f.columnconfigure((0, 1, 2), weight=1)#minsize=100
open_f.rowconfigure((0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10), weight=1)#minsize=40
filename = StringVar()
def open_file(*args):
    global filename, wb, ws, filen
    filename.set(filedialog.askopenfilename(filetypes=[('Excel File', '.xlsx')], initialdir='../'))
    b = list(filename.get())
    ind = b[-1:-(len(b)):-1].index('/')
    filen=''.join(b[-ind:-5])
    wb = load_workbook(filename.get())
    ws = wb['Sheet1']
    rowcol()
    rawt()
    if filename.get()!=None:
        open_b1['state']='!disabled'
        open_b2['state']='!disabled'
        open_b4['state']='!disabled'
def proceed(*args):
    global roll_var, warn3, warn75, y_warn_text, today_atten, deb_con, deb_low, sum_per, war_3, war_75, roll_l1, roll_l2, per, counter
    nb.tab(1, state='normal')
    #tim()
    roll_var.set(Roll)
    warn3.set('----------------------------------')
    warn75.set('----------------------------------')
    y_warn_text.set('WARNING!!!\nAbsence-?')
    today_atten = [None for i in range(tot_stud)]
    deb_con = [None for i in range(tot_stud)]
    deb_low = [None for i in range(tot_stud)]
    sum_per = [None for i in range(tot_stud)]
    war_3 = [None for i in range(tot_stud)]
    war_75 = [None for i in range(tot_stud)]
    roll_l1.set(Roll[0])
    roll_l2.set('START')
    per.set('%')
    nb.select(1)
    change_child()
    counter=0
    lbox.see(0)
def down_file(*args):
    os.system('py -2 Gdown.py')

image=Image.open('pic.jpg')
img=ImageTk.PhotoImage(image)

o_bg_l = ttk.Label(open_f, background='black')
o_bg_l.grid(row=0, column=0, sticky='nsew', rowspan=11, columnspan=3)
photo_l = ttk.Label(open_f, image=img, anchor='sw')
photo_l.grid(row=0, column=0, rowspan=11, columnspan=3, sticky='nsew')
open_b1 = ttk.Button(open_f, text='Download File', style='my.TButton', command=down_file, width=12)
open_b1.grid(row=2, column=1, sticky='ns')

open_b = ttk.Button(open_f, text='Select File', style='my.TButton', command=open_file, width=20)
open_b.grid(row=3, column=1, sticky='ns', pady=20)
open_l = ttk.Label(open_f, textvariable=filename, anchor='center', font=("Comic Sans Ms", 25, "bold"), background='red')#, padding='10'
open_l.grid(row=4, column=0, columnspan=3, sticky='nsew')
open_l1 = ttk.Label(open_f, text='*SELECT ONLY EXCEL FILES(.xlsx) CONTAINING STUDENT RECORDS\nTHE FILE SHOULD BE IN GIVEN PRE-DEFINED FORMAT', anchor='center', justify='center', font=("Times New Roman", 20, "bold"), foreground='red', background='black')
open_l1.grid(row=0, column=0, columnspan=3, sticky='nsew')
open_b1 = ttk.Button(open_f, text='SUM', command=sum_par1, style='my.TButton', state='disabled', width=30)
open_b1.grid(row=6, column=1, sticky='ns', pady=10)
open_b2 = ttk.Button(open_f, text='DEBAR', command=debar, style='my.TButton', state='disabled', width=20)
open_b2.grid(row=7, column=1, sticky='ns', pady=10)
open_b3 = ttk.Label(open_f, text='YOU CAN DIRECTLY CLICK "DO" IF ATTENDANCE SUM, % AND DEBAR LIST IS ALREADY UPDATED IN SHEET', anchor='center', font=("Times New Roman", 15, "bold"), foreground='red', background='black')
open_b3.grid(row=10, column=0, columnspan=3, sticky='nsew', pady='0 10')
open_b4 = ttk.Button(open_f, text='DO', command=proceed, style='my.TButton', state='disabled', width=10)
open_b4.grid(row=8, column=1, sticky='ns', pady=10)


#filename1=filename.get()
nb.add(open_f, text='OPEN')








mainframe = ttk.Frame(nb, padding='5 5 5 5', borderwidth=10, relief='solid')#width=1000, height=650,
mainframe.grid(row=0, column=0, sticky='nsew')
mainframe.columnconfigure((0, 1, 2), weight=1)#minsize=100
mainframe.rowconfigure((0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10), weight=1)#minsize=40





# y_warn, y_deb, d_less, d_con =[]



style = ttk.Style(mainframe)
style.configure('sl.TLabel', font=("Verdana", 25, "bold"), background='red')
style.configure('un.TLabel', font=("Verdana", 25, "bold"), background='green')
style.configure('def.TLabel', font=("Verdana", 25, "bold"), background='white')
style.configure('slw.TLabel', font=("Verdana", 25, "bold"), background='yellow')
style.configure('my.TButton', font=("Verdana", 25, "bold"))
style.configure('warel.TLabel', font=("Times New Roman", 25, "bold"), background='black', foreground='yellow')
#style.configure('newl.TCheckbutton', font=("Times New Roman", 25, "bold"), background='black', foreground='yellow')

#imag=PhotoImage(file='pic1 [Original Size].png')
photo_l1 = ttk.Label(mainframe, background='black', anchor='center')
photo_l1.grid(row=0, column=0, rowspan=11, columnspan=4, sticky='nsew')
roll_l = ttk.Label(mainframe, textvariable=roll_l1, anchor='center', justify='center', font=('Comic Sans Ms', 70, 'bold'), backgroun='black', foreground='white')
roll_l.grid(row=1, column=0, sticky='nsew', columnspan=3, rowspan=3)
roll_la = ttk.Label(mainframe, textvariable=roll_l2, anchor='center', style='def.TLabel')
roll_la.grid(row=4, column=1, sticky='nsew')
deb_lab = ttk.Label(mainframe, text='DEBAR STATUS', anchor='center', font=("Times New Roman", 25, "bold underline"), background='black', foreground='white')
deb_lab.grid(row=5, column=1, sticky='nsew')
deb_lab1 = ttk.Label(mainframe, text='ATTENDANCE\n< 75%', anchor='center', justify='center', style='def.TLabel', relief='sunken', padding='10')
deb_lab1.grid(row=6, column=0, sticky='nsew',)
deb_lab3 = ttk.Label(mainframe, textvariable=per, anchor='center', font='TkCaptionFont 45 bold', background='black', foreground='white')
deb_lab3.grid(row=6, column=1, sticky='nsew')
deb_lab2 = ttk.Label(mainframe, text='ABSENT\nFOR 5 DAYS', anchor='center', justify='center', style='def.TLabel', relief='sunken', padding='10')
deb_lab2.grid(row=6, column=2, sticky='nsew',)
war_lab1 = ttk.Label(mainframe, textvariable=warn3, anchor='center', style='warel.TLabel', padding='10')
war_lab1.grid(row=7, column=0, sticky='nsew', columnspan=3)
war_lab2 = ttk.Label(mainframe, textvariable=warn75, anchor='center', style='warel.TLabel', padding='10')
war_lab2.grid(row=8, column=0, sticky='nsew', columnspan=3)
yes_lab = ttk.Label(mainframe, text='YESTERDAY', anchor='center',font=("Times New Roman", 25, "bold underline"), background='black', foreground='white')
yes_lab.grid(row=9, column=1, sticky='nsew')
yes_lab1 = ttk.Label(mainframe, text='ATTENDANCE', anchor='center', style='def.TLabel', relief='sunken', padding='10')
yes_lab1.grid(row=10, column=0, sticky='nsew', padx=5, pady='0 10')
yes_lab2 = ttk.Label(mainframe, textvariable=y_warn_text, justify='center', anchor='center', style='def.TLabel', relief='sunken', padding='10')
yes_lab2.grid(row=10, column=1, sticky='nsew', padx=5, pady='0 10')
yes_lab3 = ttk.Label(mainframe, text='DEBARRED\n5 ABSENCE', anchor='center', justify='center', style='def.TLabel', relief='sunken', padding='10')
yes_lab3.grid(row=10, column=2, sticky='nsew', padx=5, pady='0 10')
lbox = Listbox(mainframe, listvariable=roll_var, width=7, height=10, font=('Verdana', 32, 'bold'), relief='ridge', borderwidth='10', background='black', foreground='white')
lbox.grid(row=1, column=3, rowspan=9, sticky='nsew')
s = ttk.Scrollbar(mainframe, orient=VERTICAL, command=lbox.yview)
s.grid(row=1, column=4, rowspan=9, sticky='ns')
lbox.configure(yscrollcommand=s.set)


def do_func(*args):
    global today_atten, tot_stud, deb_con, deb_low, done_debar_con, done_abs, done_debar_low, done_absdebar, done_warn, war_3, war_75
    nb.tab(2, state='normal')
    done_abs.set([Roll[i] for i in range(0, tot_stud) if today_atten[i]==0])
    done_debar_con.set([Roll[i] for i in range(0, tot_stud) if deb_con[i]==1])
    done_debar_low.set([Roll[i] for i in range(0, tot_stud) if deb_low[i]==1])
    #done_warn.set([Roll[i] for i in range(0, tot_stud) if war_3[i]==1])
    #done_absdebar.set([Roll[i] for i in range(0, tot_stud) if war_75[i]==1])
    att.set('ATTENDANCE ---- {} / {}'.format(sum(today_atten), tot_stud))
    nb.select(2)

new_but = ttk.Button(mainframe, text='DONE', command=do_func, state='disabled', style='my.TButton')
new_but.grid(row=10, column=3, sticky='nsew')

nb.add(mainframe, text='DO', state='disabled')

Broll=[]

def check_yes():
    global Broll, counter, deb_con, deb_low
    sl=counter+1
    Broll=cur_parti_Atte(sl)
    if Broll[-1]==0:
        yes_lab1.configure(style='sl.TLabel')
    else:
        yes_lab1.configure(style='un.TLabel')
    if sum(Broll[-1:-5:-1])==0:
        y_warn_text.set('WARNING!!!\nAbsence-4')
        yes_lab2.configure(style='sl.TLabel')
    elif sum(Broll[-1:-4:-1])==0:
        y_warn_text.set('WARNING!!!\nAbsence-3')
        yes_lab2.configure(style='sl.TLabel')
    else:
        y_warn_text.set('WARNING!!!\nAbsence-?')
        yes_lab2.configure(style='un.TLabel')
    if sum(Broll)==0: # add the yesterday debarred due to less than 75%
        yes_lab3.configure(style='sl.TLabel')
        yes_lab2.configure(style='un.TLabel')
    else:
        yes_lab3.configure(style='un.TLabel')
    if deb_low[counter]==None:
        deb_lab1.configure(style='def.TLabel')
    elif deb_low[counter]==0:
        deb_lab1.configure(style='un.TLabel')
    elif deb_low[counter]==1:
        deb_lab1.configure(style='sl.TLabel')
    if deb_con[counter]==None:
        deb_lab2.configure(style='def.TLabel')
    elif deb_con[counter]==0:
        deb_lab2.configure(style='un.TLabel')
    elif deb_con[counter]==1:
        deb_lab2.configure(style='sl.TLabel')
    if today_atten[counter] == None:
        roll_la.configure(style='def.TLabel')
    elif today_atten[counter]==1:
        roll_la.configure(style='un.TLabel')
    elif today_atten[counter]==0:
        roll_la.configure(style='sl.TLabel')
    if sum_per[counter]!=None:
        per.set(sum_per[counter])

    if war_3[counter]==0:
        warn3.set('----------------------------------')
        war_lab1.configure(style='warel.TLabel')
    elif war_3[counter]==1:
        warn3.set('WARNING!!! ABSENT FOR 3 DAYS')
        war_lab1.configure(style='slw.TLabel')
    elif war_3[counter]==2:
        warn3.set('WARNING!!! ABSENT FOR 4 DAYS')
        war_lab1.configure(style='slw.TLabel')
    elif war_3[counter]==3:
        warn3.set('ABSENT FOR MORE THAN 5 DAYS')
        war_lab1.configure(style='slw.TLabel')
    if war_75[counter]==1:
        warn75.set('WARNING!!! 75% < ATTENDANCE < 80%')
        war_lab2.configure(style='slw.TLabel')
    elif war_75[counter]==0:
        warn75.set('----------------------------------')
        war_lab2.configure(style='warel.TLabel')





def check_deb(sl, at):
    global Broll, counter, tcol, deb_low, deb_con, sum_per
    Broll=cur_parti_Atte(sl)
    if ws.cell(row=sl+5, column=5).value==1:
        deb_con[counter]=1
        deb_lab2.configure(style='sl.TLabel')
    elif sum(Broll[1:5])+at==0:
        q=messagebox.askyesno(message='ALERT!!! Absent for 5 days Continuosly\nWant to DEBAR?', icon='question', title='DEBAR ALERT')
        deb_con[counter]=int(q)#Add 2 for today debarred student
        if q:
            deb_lab2.configure(style='sl.TLabel')
        else:
            deb_lab2.configure(style='un.TLabel')
    else:
        deb_con[counter]=0
        deb_lab2.configure(style='un.TLabel')
    if ws.cell(row=sl+5, column=4 ).value==1:
        if (ws.cell(row=sl+5, column=6).value+at)/(tcol-6)*100>75:
            messagebox.showinfo(message='CONGO!!! Your Attendance > 75% now\nYou have been cleared from DEBAR LIST')
            deb_low[counter]=0
        else:
            deb_low[counter]=1
            deb_lab1.configure(style='sl.TLabel')
    else:
        if (ws.cell(row=sl+5, column=6).value+at)/(tcol-6)*100<75:
            p=messagebox.askyesno(message='ALERT!!! Attendance < 75%\nWant to DEBAR?', icon='question', title='DEBAR ALERT')
            deb_low[counter]=int(p)#Add 2 for today debarred student
            if p:
                deb_lab1.configure(style='sl.TLabel')
            else:
                deb_lab1.configure(style='un.TLabel')
        else:
            deb_low[counter]=0
            deb_lab1.configure(style='un.TLabel')
    sum_per[counter] = str(round((ws.cell(row=sl+5, column=6).value+at)/(tcol-6)*100, 2))+'%'
    per.set(sum_per[counter])
    if sum(Broll)+at==0:
        warn3.set('ABSENT FOR MORE THAN 5 DAYS')
        war_lab1.configure(style='slw.TLabel')
        war_3[counter]=3
    elif sum(Broll[2:5])+at==0:
        warn3.set('WARNING!!! ABSENT FOR 4 DAYS')
        war_lab1.configure(style='slw.TLabel')
        war_3[counter]=2
    elif sum(Broll[3:5])+at==0:
        warn3.set('WARNING!!! ABSENT FOR 3 DAYS')
        war_lab1.configure(style='slw.TLabel')
        war_3[counter]=1
    else:
        warn3.set('----------------------------------')
        war_lab1.configure(style='warel.TLabel')
        war_3[counter]=0
    if 75<(ws.cell(row=sl+5, column=6).value+at)/(tcol-6)*100<80:
        warn75.set('WARNING!!! 75% < ATTENDANCE < 80%')
        war_lab2.configure(style='slw.TLabel')
        war_75[counter]=1
    else:
        warn75.set('----------------------------------')
        war_lab2.configure(style='warel.TLabel')
        war_75[counter]=0

def change_child(*args):
    global tot_stud
    roll_la.configure(style='def.TLabel')
    deb_lab2.configure(style='def.TLabel')
    deb_lab1.configure(style='def.TLabel')
    war_lab1.configure(style='warel.TLabel')
    war_lab2.configure(style='warel.TLabel')
    yes_lab1.configure(style='def.TLabel')
    yes_lab2.configure(style='def.TLabel')
    yes_lab3.configure(style='def.TLabel')
    for i in range(0, tot_stud):
        lbox.itemconfigure(i, background='black', foreground='white')
    #lbox.configure(style='def.TLabel')
    #for child in mainframe.winfo_children():
     #   child.configure(style='def.TLabel')
'''
def do_func(*args):
    global today_atten, tot_stud, deb_con, deb_low, done_debar_con, done_abs, done_debar_low, done_absdebar, done_warn, war_3, war_75
    done_abs.set([Roll[i] for i in range(0, tot_stud) if today_atten[i]==0])
    done_debar_con.set([Roll[i] for i in range(0, tot_stud) if deb_con[i]==1])
    done_debar_low.set([Roll[i] for i in range(0, tot_stud) if deb_low[i]==1])
    done_warn.set([Roll[i] for i in range(0, tot_stud) if war_3[i]==1])
    done_absdebar.set([Roll[i] for i in range(0, tot_stud) if war_75[i]==1])
'''
def save_f(*args):
    global trow, tcol, today_atten, deb_con, deb_low
    j=0
    for i in range(6, trow+1):
        ws.cell(row=i, column=tcol+1).value = today_atten[j]
        j+=1
    tcol+=1
    sum_par()
    tcol-=1
    for i in range(1, trow-4):
        ws.cell(row=i+5, column=5).fill=whiteFill
        ws.cell(row=i+5, column=4).fill=whiteFill
        if deb_low[i-1]==1:
            ws.cell(row=i+5, column=4).value=1
            ws.cell(row=i+5, column=4).fill=redFill
        if deb_con[i-1]==1:
            ws.cell(row=i+5, column=5).value=1
            ws.cell(row=i+5, column=5).fill=yellowFill
    tim()
    q=messagebox.askyesno(message='The Attendance Sheet is going to be updated as a new file named\n\n{}_ODD.xlsx\n\nAre you sure you want to continue?'.format(filen), icon='question', title='Attendance Sheet Updation')
    if int(q)==1:
        wb.save('{}_ODD.xlsx'.format(filen))


def uplo(*args):
    global filen
    save_f()
    os.system('py -2 Gup.py')

def send_sms(*args):
    mom=messagebox.askyesno(message='Are you sure you want to send the SMS?', icon='question', title='SENDING SMS!!!')
    if int(mom)==1:
        global check1, check2, wb, today_atten, deb_con, deb_low, t
        accountSID = 'ACc1594b52b30ea17b53e87812802b4403'
        authToken = 'cbf5de4615574ef994c962e5ae710530'
        twilioNumber = '+17865040726'
        ws1 = wb['Sheet2']
        message = t.get("1.0", 'end-1c')
        def textmyself(message, rnum):
            twilioCli = TwilioRestClient(accountSID, authToken)
            twilioCli.messages.create(body=message, from_=twilioNumber, to=rnum)
            print('SMS SENT')

        if check1.get()==1:
            ab = [ ws1.cell(row=i+3, column=4).value for i in range(0, tot_stud) if (today_atten[i]==0 and ws1.cell(row=i+3, column=4).value!=None)]
            for num in ab:
                textmyself(message, '+91'+str(num))
            print(ab)

        if check2.get()==1:
            ab1 = [ ws1.cell(row=i+3, column=4).value for i in range(0, tot_stud) if (deb_con[i]==1 and ws1.cell(row=i+3, column=4).value!=None)]
            for num in ab1:
                textmyself(message, '+91'+str(num))
            print(ab1)
            ab2 = [ ws1.cell(row=i+3, column=4).value for i in range(0, tot_stud) if (deb_low[i]==1 and ws1.cell(row=i+3, column=4).value!=None)]
            for num in ab2:
                textmyself(message, '+91'+str(num))
            print(ab2)

def send_mail(*args):
    mom = messagebox.askyesno(message='Are you sure you want to send the E-MAIL?', icon='question', title='SENDING E-MAIL!!!')
    if int(mom)==1:
        pass

today_atten=[]#remove this
tot_stud=0#remove this

image1=Image.open('pic1.jpg')
img1=ImageTk.PhotoImage(image1)




check1 = IntVar()
check2 = IntVar()
doneframe = ttk.Frame(nb, padding='5 5 5 5', borderwidth=10, relief='solid')#width=1000, height=650,
doneframe.grid(row=0, column=0, sticky='nsew')
doneframe.columnconfigure((6, 7), weight=1)#minsize=100
doneframe.rowconfigure((0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14), weight=1)#minsize=40
d_bg_l = ttk.Label(doneframe, background='black')
d_bg_l.grid(row=0, column=0, sticky='nsew', rowspan=15, columnspan=8)
photo_l = ttk.Label(doneframe, image=img1, anchor='sw')
photo_l.grid(row=0, column=0, rowspan=15, columnspan=8, sticky='nsew')
at_l = ttk.Label(doneframe, textvariable=att, anchor='center', font=('Verdana', 50, 'bold'), backgroun='black', foreground='#00FF00')
at_l.grid(row=0, column=0, columnspan=8, sticky='nsew')
d_abs = ttk.Label(doneframe, text='ABSENTEES', anchor='center', font=("Times New Roman", 30, "bold underline"), background='black', foreground='white')
d_abs.grid(row=2, column=0, columnspan=2, sticky='nsew')
d_lbox2 = Listbox(doneframe, listvariable=done_abs, width=7, height=10, font=('Verdana', 25, 'bold'), relief='ridge', borderwidth='10', background='red')
d_lbox2.grid(row=4, column=0, rowspan=10, sticky='ns', padx='40 0', pady='0 40')
d_s1 = ttk.Scrollbar(doneframe, orient=VERTICAL, command=d_lbox2.yview)
d_s1.grid(row=4, column=1, rowspan=10, sticky='ns', padx='0 40', pady='0 40')
d_lbox2.configure(yscrollcommand=d_s1.set)
d_deb_l = ttk.Label(doneframe, text='DEBARRED LIST',anchor='center', font=("Times New Roman", 30, "bold underline"), background='black', foreground='white')
d_deb_l.grid(row=2, column=2, columnspan=4, sticky='nsew')
d_deb_l1 = ttk.Label(doneframe, text='< 75%', anchor='center', font=("Times New Roman", 25, "bold underline"), background='black', foreground='red')
d_deb_l1.grid(row=3, column=2, columnspan=2, sticky='nsew')
d_deb_l2 = ttk.Label(doneframe, text='5 Absent', anchor='center', font=("Times New Roman", 25, "bold underline"), background='black',  foreground='red')
d_deb_l2.grid(row=3, column=4, columnspan=2, sticky='nsew')
d_lbox = Listbox(doneframe, listvariable=done_debar_low, width=7, height=10, font=("Verdana", 25, "bold"), relief='ridge', borderwidth='10', background='red')
d_lbox.grid(row=4, column=2, rowspan=10, sticky='ns', pady='0 40')
d_s2 = ttk.Scrollbar(doneframe, orient=VERTICAL, command=d_lbox.yview)
d_s2.grid(row=4, column=3, rowspan=10, sticky='ns', pady='0 40')
d_lbox.configure(yscrollcommand=d_s2.set)
d_lbox1 = Listbox(doneframe, listvariable=done_debar_con, width=7, height=10, font=("Verdana", 25, "bold"), relief='ridge', borderwidth='10', background='red')
d_lbox1.grid(row=4, column=4, rowspan=10, sticky='ns', pady='0 40')
d_s3 = ttk.Scrollbar(doneframe, orient=VERTICAL, command=d_lbox1.yview)
d_s3.grid(row=4, column=5, rowspan=10, sticky='ns', pady='0 40')
d_lbox1.configure(yscrollcommand=d_s3.set)
d_but = ttk.Button(doneframe, text='SAVE', command=save_f, style='my.TButton')
d_but.grid(row=12, column=7, sticky='nsew', padx='0 40')
d_not = ttk.Label(doneframe, text='NOTIFY', anchor='center', font=("Times New Roman", 30, "bold underline"), background='black', foreground='white')
d_not.grid(row=2, column=6, columnspan=2, sticky='nsew', padx='40 40')
d_check1 = Checkbutton(doneframe, text='Absentees', variable=check1, anchor='w', font=("Times New Roman", 20, 'bold'), activebackground='yellow')#, font=("Times New Roman", 25), background='red', foreground='white')
d_check1.grid(row=3, column=6, sticky='nsew', padx='40 0')
d_check2 = Checkbutton(doneframe, text='Debarred', variable=check2, anchor='w', font=("Times New Roman", 20, 'bold'), activebackground='yellow')#, , background='black', foreground='white')
d_check2.grid(row=3, column=7, sticky='nsew', padx='0 40')
'''d_check3 = Checkbutton(doneframe, text='Absentees', variable=check3, anchor='w', font=("Times New Roman", 20, 'bold'), activebackground='yellow')#, font=("Times New Roman", 25), background='red', foreground='white')
d_check3.grid(row=4, column=6, sticky='ew', padx='40 0' )
d_check4 = Checkbutton(doneframe, text='Debarred', variable=check4, anchor='w', font=("Times New Roman", 20, 'bold'),activebackground='yellow')#, , background='black', foreground='white')
d_check4.grid(row=4, column=7, sticky='ew', padx='0 40' )'''
t = Text(doneframe, width=20, height=6, wrap='word', bd=10, font=('Times New Roman', 25, 'bold'), bg='black', fg='yellow', insertbackground='white')
t.grid(row=4, column=6, rowspan=6, columnspan=2, sticky='nsew', padx='40 40')
#d_but1 = ttk.Button(doneframe, text='SEND', command=uplo, style='my.TButton')
#d_but1.grid(row=11, column=6, columnspan=2, sticky='nsew', padx='40 40')
d_but2 = ttk.Button(doneframe, text='UPLOAD', command=uplo, style='my.TButton')
d_but2.grid(row=12, column=6, sticky='nsew', padx='40 0')
d_but1 = ttk.Button(doneframe, text='SMS', command=send_sms, style='my.TButton')
d_but1.grid(row=10, column=6, sticky='nsew', padx='40 0')
d_but2 = ttk.Button(doneframe, text='E-mail', command=send_mail, style='my.TButton')
d_but2.grid(row=10, column=7, sticky='nsew', padx='0 40')



nb.add(doneframe, text='DONE', state='disabled')

def right(*args):
    global counter, today_atten
    if (-1)<counter<tot_stud:
        lbox.see(counter)
        today_atten[counter]=1
        lbox.itemconfigure(counter, background='green', foreground='black')
        check_deb(counter+1,1)
        check_yes()
        if counter==tot_stud-1:
            roll_l1.set('FINISHED')
            roll_l2.set(Roll[counter])
            new_but['state']='!disabled'
        else:
            counter+=1
            roll_l1.set(Roll[counter])
            roll_l2.set(Roll[counter-1])
    #lbox.see(counter)

def left(*args):
    global counter, today_atten
    if (-1)<counter<tot_stud:
        lbox.see(counter)
        today_atten[counter]=0
        lbox.itemconfigure(counter, background='red', foreground='black')
        check_deb(counter+1, 0)
        check_yes()
        if counter==tot_stud-1:
            roll_l1.set('FINISHED')
            roll_l2.set(Roll[counter])
            new_but['state']='!disabled'
        else:
            counter+=1
            roll_l1.set(Roll[counter])
            roll_l2.set(Roll[counter-1])
    #lbox.see(counter)
def up(*args):
    global counter
    if roll_l1.get()=='FINISHED':
        lbox.see(counter)
        roll_l1.set(Roll[counter])
        roll_l2.set(Roll[counter-1])
        counter-=1
        check_yes()
        counter+=1
    elif 0<counter<tot_stud:
        #lbox.see(counter)
        counter-=1
        roll_l1.set(Roll[counter])
        if counter==0:
            roll_l2.set('START')
            warn3.set('----------------------------------')
            warn75.set('----------------------------------')
            per.set('%')
            #change_child()
            roll_la.configure(style='def.TLabel')
            deb_lab2.configure(style='def.TLabel')
            deb_lab1.configure(style='def.TLabel')
            war_lab1.configure(style='warel.TLabel')
            war_lab2.configure(style='warel.TLabel')
            yes_lab1.configure(style='def.TLabel')
            yes_lab2.configure(style='def.TLabel')
            yes_lab3.configure(style='def.TLabel')
        else:
            roll_l2.set(Roll[counter-1])
            counter-=1
            check_yes()
            counter+=1
        lbox.see(counter-1)

def down(*args):
    global counter
    if (-1)<counter<tot_stud and today_atten[counter]!=None:
        lbox.see(counter)
        counter+=1
        if counter==tot_stud:
            counter-=1
            roll_l2.set(Roll[counter])
            roll_l1.set('FINISHED')
            check_yes()
        else:
            roll_l1.set(Roll[counter])
            roll_l2.set(Roll[counter-1])
            counter-=1
            check_yes()
            counter+=1

def dummy(*args):
    open_b1['state']='disabled'
    new_but['state']='disabled'



#roll_l1.trace('w', check_yes)
root.bind('<Right>', right)
root.bind('<Left>', left)
root.bind('<Up>', up)
root.bind('<Down>', down)
#nb.tab(1).bind('<1>', dummy)


root.mainloop()

#wb.save('new.xlsx')























